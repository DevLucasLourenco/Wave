import json
import os
import time
import openpyxl
import pandas as pd

class Archive:
    
    def __init__(self, file) -> None:
        self.__DesignatedFile = file
        self.__DictWithData:dict[int|str] = dict()
        self.__Dataframe:pd.DataFrame
        self.__Filename:str
        self.__FileType:str
        self.__FileMetaData:dict
        self.__background_run()


    def __background_run(self):
        self.__type_file()
        self.__name_of_file()
        self.__metadata_of_file()


    def __name_of_file(self):
        self.__Filename = str(self.__DesignatedFile).replace(self.__FileType, "")[:-1]
    

    def __type_file(self):
        self.__FileType = self.__DesignatedFile.split('.')[-1]
    

    def __metadata_of_file(self):    
        file_size = os.path.getsize(self.__DesignatedFile)
        file_name = os.path.basename(self.__DesignatedFile)
        last_modified = time.ctime(os.path.getmtime(self.__DesignatedFile))
        file_permissions = oct(os.stat(self.__DesignatedFile).st_mode)[-3:]

        metadata = {
            "file_name": file_name,
            "file_size": f"{file_size} bytes",
            "file_type": self.__FileType,
            "last_modified": last_modified,
            "file_permissions": file_permissions}
        
        match self.__FileType:
            case 'csv':
                df = pd.read_csv(self.__DesignatedFile)
                metadata.update({
                    "columns": df.columns.tolist(),
                    "num_rows": df.shape[0],
                    "num_columns": df.shape[1]
                }
            )

            case 'xlsx':
                wb = openpyxl.load_workbook(self.__DesignatedFile)
                metadata.update({
                    "num_sheets": len(wb.sheetnames),
                    "sheet_names": wb.sheetnames,
                }
            )
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    metadata.update({
                        f"{sheet}_dimensions": ws.dimensions,
                        f"{sheet}_max_rows": ws.max_row,
                        f"{sheet}_max_columns": ws.max_column
                    }
                )

            case 'json':
                with open(self.__DesignatedFile, 'r') as file:
                    data = json.load(file)
                metadata.update({
                    "num_keys": len(data),
                    "keys": list(data.keys())
                }
            )
    
            case _:
                pass
            
        self.__FileMetaData = metadata
    

    def __turnDataframeIntoDict(self):
        self.__DictWithData = self.__Dataframe.to_dict(orient='list')
        self.__refactoringDictToPatterns()


    def __refactoringDictToPatterns(self):
        for k, v in self.__DictWithData.items():
            typeOfValue = type(v[0])
            self.__DictWithData.update({k:{"type_column":typeOfValue, 'data_column':v}})
       

    def getFileType(self) -> str:
        return self.__FileType


    def getMetaData(self) -> list[str]:
        return self.__FileMetaData


    def getFilename(self) -> str:
        return self.__Filename
    

    def getDesignatedFile(self):
        return self.__DesignatedFile
    

    def getData(self) -> dict[int, str]:
        if self.__DictWithData:
            return self.__DictWithData
        raise ReferenceError('No Dict Available')
    

    def getDataFrame(self) -> pd.DataFrame:
        if not self.__Dataframe.empty:
            return self.__Dataframe
        raise ReferenceError('No DataFrame Available')
    

    def setDataframe(self, df):
        self.__Dataframe = df
        self.__turnDataframeIntoDict()
